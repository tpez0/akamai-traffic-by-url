#!/usr/bin/env python3
"""
CLI per interrogare Akamai Reporting API (urlbytes-by-url) da shell.

Aggiornamenti:
  - `--limit MAX`: split per giorno (UTC) e download di 25.000 record/giorno
    con append su CSV/XLSX e colonna `day` sempre presente.
  - Supporto **Excel**: `--format xlsx` (usa openpyxl).
  - Log finale dei giorni senza dati.
  - Messaggi di avanzamento per ogni giorno: "Download YYYY-MM-DD...".
  - Parser robusto del payload (columns+rows, data[], metric -> {url -> valore}).
  - Modalità interattiva (`--interactive`) e opzioni di troubleshooting
    (`--dry-run`, `-v`, `--log-headers`, `--timeout`).

Installazione:
  pip install requests edgegrid-python openpyxl
  Configura ~/.edgerc con credenziali EdgeGrid, es.:
    [default]
    client_token = akab-...
    client_secret = ...
    access_token = akab-...
    host = https://{tuo-host}.luna.akamaiapis.net

Esempio (MAX su luglio, XLSX):
  python report_urlbytes.py \
    --start 2025-07-01T00:00:00Z \
    --end   2025-08-01T00:00:00Z \
    --interval DAY \
    --object-id 1836353 --object-id 1508185 \
    --metric allEdgeBytes --metric allOriginBytes --metric allBytesOffload \
    --limit MAX --format xlsx --out report_july.xlsx -v
"""
from __future__ import annotations
import argparse
import configparser
import csv
import json
import os
import sys
import time
from datetime import datetime, timedelta, timezone
from typing import List, Optional, Dict, Any

import requests
try:
    from akamai.edgegrid import EdgeGridAuth
except Exception as e:  # pragma: no cover
    eprint("[ERRORE] Manca la libreria 'edgegrid-python' (modulo akamai.edgegrid).")
    eprint("Installa con: pip install requests edgegrid-python")
    sys.stderr.write(f"Dettagli: {e}")
    sys.stderr.flush()
    sys.exit(2)

# Excel (openpyxl è opzionale, richiesto solo con --format xlsx)
try:
    from openpyxl import Workbook, load_workbook  # type: ignore
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False

def eprint(*args, **kwargs):
    """Print su stderr con newline automatico."""
    print(*args, file=sys.stderr, **kwargs)

DEFAULT_REPORT = "urlbytes-by-url"
DEFAULT_VERSION = 1

SUGGESTED = {
    "start": os.getenv("REPORT_START", "2025-07-01T00:00:00Z"),
    "end": os.getenv("REPORT_END", "2025-08-01T00:00:00Z"),
    "interval": os.getenv("REPORT_INTERVAL", "DAY"),
    "object_type": os.getenv("REPORT_OBJECT_TYPE", "cpcode"),
    "object_ids": os.getenv("REPORT_OBJECT_IDS", "1836353,1508185"),
    "metrics": os.getenv("REPORT_METRICS", "allEdgeBytes,allOriginBytes,allBytesOffload"),
    "limit": os.getenv("REPORT_LIMIT", "5000"),  # string: numero oppure 'MAX'
}

# -------------------- util --------------------

def read_edgerc(path: str, section: str) -> dict:
    cfg = configparser.ConfigParser()
    if not cfg.read(path):
        raise FileNotFoundError(f"Impossibile leggere {path}")
    if section not in cfg:
        raise KeyError(f"Sezione '{section}' non trovata in {path}")
    sec = cfg[section]
    host = sec.get("host", fallback=None)
    if not host:
        raise KeyError("Chiave 'host' mancante nella sezione dell'edgerc")
    host = host.strip()
    if not host.startswith("http://") and not host.startswith("https://"):
        host = "https://" + host
    host = host.rstrip("/")
    return {
        "host": host,
        "client_token": sec.get("client_token"),
        "client_secret": sec.get("client_secret"),
        "access_token": sec.get("access_token"),
    }


def build_session(creds: dict) -> requests.Session:
    s = requests.Session()
    s.auth = EdgeGridAuth(
        client_token=creds["client_token"],
        client_secret=creds["client_secret"],
        access_token=creds["access_token"],
    )
    return s


def _parse_iso_z(s: str) -> datetime:
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _iso_z(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

def _day_to_range(day: str) -> tuple[str, str]:
    """Ritorna (startZ, endZ) per il giorno in formato YYYY-MM-DD."""
    d = datetime.strptime(day, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    start = _iso_z(d)
    end = _iso_z(d + timedelta(days=1))
    return start, end


# -------------------- parsing payload → records --------------------

def _flatten_dict(d: Dict[str, Any], parent: str = "", sep: str = ".") -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for k, v in d.items():
        key = f"{parent}{sep}{k}" if parent else k
        if isinstance(v, dict):
            out.update(_flatten_dict(v, key, sep))
        else:
            out[key] = v
    return out


def _records_from_response(payload: Any, metrics_hint: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    """Trasforma risposte comuni Akamai in righe (dict) pronte per CSV/XLSX.
    Supporta:
      • schema columns+rows
      • schema data[] con dimensions/metrics
      • schema { metric → { url → valore } } (con metrics_hint)
    """
    records: List[Dict[str, Any]] = []

    if isinstance(payload, dict):
        # columns + rows
        if isinstance(payload.get("columns"), list) and isinstance(payload.get("rows"), list):
            cols: List[str] = []
            for c in payload["columns"]:
                cols.append(str(c.get("name") if isinstance(c, dict) else c))
            for row in payload["rows"]:
                if isinstance(row, list):
                    rec = {cols[i] if i < len(cols) else f"col{i}": row[i] for i in range(len(row))}
                    records.append(rec)
                elif isinstance(row, dict):
                    records.append(row)
            return records

        # data: [ ... ]
        if isinstance(payload.get("data"), list):
            for item in payload["data"]:
                if isinstance(item, dict):
                    rec: Dict[str, Any] = {}
                    dims = item.get("dimensions") if isinstance(item.get("dimensions"), dict) else None
                    mets = item.get("metrics") if isinstance(item.get("metrics"), dict) else None
                    if dims:
                        rec.update(dims)
                    if mets:
                        rec.update(mets)
                    for k, v in item.items():
                        if k in ("dimensions", "metrics"):
                            continue
                        if not isinstance(v, (dict, list)):
                            rec[k] = v
                    if not rec:
                        rec = _flatten_dict(item)
                    records.append(rec)
            return records

        # data: { metric -> { url -> value } } oppure top-level metrics
        data_obj = payload.get("data") if isinstance(payload.get("data"), dict) else None
        metric_map = payload if (metrics_hint and all(k in payload for k in metrics_hint)) else data_obj
        if metrics_hint and isinstance(metric_map, dict) and all(isinstance(metric_map.get(m), dict) for m in metrics_hint):
            urls = set()
            for m in metrics_hint:
                urls.update(metric_map.get(m, {}).keys())
            for u in sorted(urls):
                rec: Dict[str, Any] = {"url": u}
                for m in metrics_hint:
                    val = metric_map.get(m, {}).get(u)
                    if val is not None and not isinstance(val, (dict, list)):
                        rec[m] = val
                records.append(rec)
            return records

        return [_flatten_dict(payload)]

    if isinstance(payload, list):
        for it in payload:
            records.append(_flatten_dict(it) if isinstance(it, dict) else {"value": it})
        return records

    return [{"value": payload}]


# -------------------- CSV helpers --------------------

def _write_csv_header(out_path: Optional[str], keys: List[str], delimiter: str) -> None:
    f = open(out_path, "w", encoding="utf-8-sig", newline="") if out_path else sys.stdout
    writer = csv.DictWriter(f, fieldnames=keys, delimiter=delimiter, lineterminator='\r\n')
    writer.writeheader()
    if out_path:
        f.close()


def _append_csv_rows(out_path: Optional[str], keys: List[str], rows: List[Dict[str, Any]], delimiter: str) -> None:
    f = open(out_path, "a", encoding="utf-8-sig", newline="") if out_path else sys.stdout
    writer = csv.DictWriter(f, fieldnames=keys, delimiter=delimiter, lineterminator='\r\n')
    for r in rows:
        writer.writerow({k: r.get(k, "") for k in keys})
    if out_path:
        f.close()


# -------------------- XLSX helpers --------------------

def _ensure_xlsx_available():
    if not HAVE_OPENPYXL:
        eprint("[ERRORE] Per --format xlsx serve 'openpyxl'. Installa con: pip install openpyxl")
        sys.stderr.flush()
        sys.exit(2)


def _write_xlsx_header(out_path: str, keys: List[str], sheet_name: str) -> None:
    _ensure_xlsx_available()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(keys)
    wb.save(out_path)


def _append_xlsx_rows(out_path: str, keys: List[str], rows: List[Dict[str, Any]], sheet_name: str) -> None:
    _ensure_xlsx_available()
    try:
        wb = load_workbook(out_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(keys)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(keys)
    for r in rows:
        ws.append([r.get(k, "") for k in keys])
    wb.save(out_path)

# -------------------- hostname helpers --------------------

def _extract_hostname(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    v = value.strip()
    if not v:
        return ""
    # Gestione URL con schema
    if "://" in v:
        try:
            from urllib.parse import urlparse
            return urlparse(v).netloc
        except Exception:
            pass
    # Rimuovi eventuale prefisso "//"
    if v.startswith("//"):
        v = v[2:]
    return v.split("/", 1)[0]


def _ensure_hostname_column(records: List[Dict[str, Any]]) -> None:
    """Aggiunge la colonna 'hostname' se manca, ricavandola da 'url' o 'hostname.url'."""
    for rec in records:
        if rec.get("hostname"):
            continue
        candidate = None
        for key in ("hostname.url", "url", "request.url"):
            val = rec.get(key)
            if isinstance(val, str) and val:
                candidate = val
                break
        if candidate:
            rec["hostname"] = _extract_hostname(candidate)


# -------------------- Modalità interattiva --------------------

def _ask(prompt: str, default: Optional[str] = None) -> str:
    msg = f"{prompt} [{default}]: " if default is not None else f"{prompt}: "
    val = input(msg).strip()
    return val or (default or "")


def interactive_fill(args: argparse.Namespace) -> None:
    if not args.start:
        args.start = _ask("Start (ISO8601 Z)", SUGGESTED["start"])
    if not args.end:
        args.end = _ask("End (ISO8601 Z)", SUGGESTED["end"])
    if not args.interval:
        args.interval = _ask("Interval (HOUR|DAY|WEEK|MONTH)", SUGGESTED["interval"]).upper()
    if not args.object_type:
        args.object_type = _ask("Object type", SUGGESTED["object_type"]) 

    if not args.object_ids:
        ids = _ask("Object IDs (comma-separati)", SUGGESTED["object_ids"]) 
        args.object_ids = [x.strip() for x in ids.split(",") if x.strip()]

    if not args.metrics:
        mets = _ask("Metrics (comma-separate)", SUGGESTED["metrics"]) 
        args.metrics = [x.strip() for x in mets.split(",") if x.strip()]

    if args.limit is None:
        lim = _ask("Limit (numero o MAX)", str(SUGGESTED["limit"]))
        args.limit = lim or SUGGESTED["limit"]

    if not args.format:
        fmt = _ask("Output format (json/csv/xlsx)", "csv").lower()
        args.format = fmt if fmt in ("json", "csv", "xlsx") else "csv"
        if args.format in ("csv", "xlsx") and not args.out:
            suggested = "report.xlsx" if args.format == "xlsx" else "report.csv"
            maybe_out = _ask(f"Nome file {args.format.upper()} (vuoto={'stdout' if args.format=='csv' else 'report.xlsx'})", suggested)
            if maybe_out:
                args.out = maybe_out
            elif args.format == "xlsx":
                args.out = suggested

    if args.pretty is False and _ask("Pretty print JSON? (y/N)", "N").lower().startswith("y"):
        args.pretty = True
    if not args.verbose and _ask("Verbose? (y/N)", "N").lower().startswith("y"):
        args.verbose = True
    if not args.log_headers and _ask("Log headers? (y/N)", "N").lower().startswith("y"):
        args.log_headers = True


# -------------------- main --------------------

def main():
    parser = argparse.ArgumentParser(description="CLI per Akamai Reporting API — urlbytes-by-url")
    parser.add_argument("--start", help="ISO8601 con Z, es. 2025-07-01T00:00:00Z")
    parser.add_argument("--end", help="ISO8601 con Z, es. 2025-08-01T00:00:00Z")
    parser.add_argument("--interval", choices=["HOUR", "DAY", "WEEK", "MONTH"], help="Intervallo di aggregazione")
    parser.add_argument("--object-type", dest="object_type", help="Tipo oggetto: cpcode, property, ...")
    parser.add_argument("--object-id", dest="object_ids", action="append", help="ID oggetto (ripetibile)")
    parser.add_argument("--metric", dest="metrics", action="append", help="Nome metrica (ripetibile)")
    parser.add_argument("--limit", help="Limite record per pagina (numero) oppure 'MAX' per 25k/giorno con append CSV/XLSX")

    parser.add_argument("--report", default=DEFAULT_REPORT, help="Nome report (default urlbytes-by-url)")
    parser.add_argument("--version", type=int, default=DEFAULT_VERSION, help="Versione report (default 1)")
    parser.add_argument("--edgerc", default=os.path.expanduser(".edgerc"), help="Percorso edgerc")
    parser.add_argument("--edgerc-section", default="default", help="Sezione del file edgerc")
    parser.add_argument("--out", help="Salva la risposta su file (JSON/CSV/XLSX a seconda del formato)")

    parser.add_argument("--pretty", action="store_true", help="Stampa JSON formattato")
    parser.add_argument("--verbose", "-v", action="store_true", help="Log esteso (status, request-id)")
    parser.add_argument("--log-headers", action="store_true", help="Mostra anche le response headers")
    parser.add_argument("--dry-run", action="store_true", help="Mostra richiesta (URL, params, body) e non invia")
    parser.add_argument("--timeout", type=int, default=120, help="Timeout della richiesta in secondi (default 120)")
    parser.add_argument("--interactive", action="store_true", help="Chiedi i parametri mancanti da stdin")

    parser.add_argument("--format", choices=["json", "csv", "xlsx"], help="Formato di output (default json)")
    parser.add_argument("--csv-delimiter", default=",", help="Delimitatore CSV (default ',')")
    parser.add_argument("--sheet-name", default="Report", help="Nome del foglio Excel quando --format xlsx")
    parser.add_argument("--retry-no-data", type=int, default=10, help="Ritentativi per giorni senza dati (default 10)")
    parser.add_argument("--retry-wait", type=float, default=2.0, help="Attesa (s) tra ritentativi su giorni senza dati (default 2.0)")

    args = parser.parse_args()

    need_core = not (args.start and args.end and args.interval and args.object_type and args.object_ids and args.metrics)
    if args.interactive or need_core or not args.format:
        interactive_fill(args)

    args.start = args.start or SUGGESTED["start"]
    args.end = args.end or SUGGESTED["end"]
    args.interval = (args.interval or SUGGESTED["interval"]).upper()
    args.object_type = args.object_type or SUGGESTED["object_type"]
    args.object_ids = args.object_ids or [x.strip() for x in SUGGESTED["object_ids"].split(",") if x.strip()]
    args.metrics = args.metrics or [x.strip() for x in SUGGESTED["metrics"].split(",") if x.strip()]
    args.limit = args.limit if args.limit is not None else SUGGESTED["limit"]
    args.format = args.format or "json"

    # Default file per XLSX
    if args.format == "xlsx" and not args.out:
        args.out = "report.xlsx"

    try:
        creds = read_edgerc(args.edgerc, args.edgerc_section)
    except Exception as e:
        eprint(f"[ERRORE] Configurazione edgerc: {e}")
        sys.stderr.flush()
        sys.exit(2)

    if args.format == "xlsx" and not HAVE_OPENPYXL:
        eprint("[ERRORE] --format xlsx richiede 'openpyxl'. Installa con: pip install openpyxl")
        sys.stderr.flush()
        sys.exit(2)

    session = build_session(creds)
    base_url = f"{creds['host']}/reporting-api/v1/reports/{args.report}/versions/{args.version}/report-data"

    limit_str = str(args.limit).strip().upper() if isinstance(args.limit, str) else str(args.limit)
    max_mode = (limit_str == "MAX")
    # In precedenza forzavamo CSV; ora permettiamo CSV o XLSX. Se JSON in MAX, avviso.
    if max_mode and args.format == "json":
        eprint("[INFO] Modalità MAX: meglio usare --format csv o --format xlsx. Proseguo con JSON aggregato.")
        sys.stderr.flush()

    # --- Percorso MAX: split per-giorno ---
    if max_mode:
        args.interval = "DAY"
        start_dt = _parse_iso_z(args.start)
        end_dt = _parse_iso_z(args.end)
        cur = datetime(start_dt.year, start_dt.month, start_dt.day, tzinfo=timezone.utc)
        end_day = datetime(end_dt.year, end_dt.month, end_dt.day, tzinfo=timezone.utc)
        if end_dt > end_day:
            end_day += timedelta(days=1)

        header_keys: Optional[List[str]] = None
        total_rows = 0
        days_no_data: List[str] = []
        json_aggregate: List[Any] = []

        while cur < end_day:
            nxt = min(cur + timedelta(days=1), end_day)
            params = {"start": _iso_z(cur), "end": _iso_z(nxt), "interval": "DAY"}
            body = {
                "objectType": args.object_type,
                "objectIds": args.object_ids,
                "metrics": args.metrics,
                "limit": 25000,
            }

            print(f"Download {params['start'][:10]}...", file=sys.stderr, flush=True)

            if args.dry_run:
                cur = nxt
                continue

            try:
                r = session.post(base_url, params=params, json=body, headers={"Content-Type": "application/json"}, timeout=args.timeout)
            except requests.exceptions.Timeout:
                eprint(f"[ERRORE] Timeout {args.timeout}s per {params['start']} - {params['end']}")
                sys.stderr.flush()
                sys.exit(2)
            except requests.exceptions.RequestException as e:
                eprint(f"[ERRORE] Richiesta fallita per {params['start']} - {params['end']}: {e}")
                sys.stderr.flush()
                sys.exit(2)

            if r.status_code >= 400:
                try:
                    payload = r.json()
                except Exception:
                    payload = {"error": r.text}
                eprint(json.dumps(payload, ensure_ascii=False, indent=2))
                sys.stderr.flush()
                sys.exit(1)

            if not r.content or r.text.strip() == "":
                eprint(f"  Nessun contenuto per {params['start'][:10]}")
                sys.stderr.flush()
                days_no_data.append(params['start'][:10])
                cur = nxt
                continue

            try:
                data = r.json()
            except Exception:
                eprint(f"[ERRORE] Payload non-JSON per {params['start']} — {params['end']}.")
                sys.stderr.flush()
                sys.exit(1)

            recs = _records_from_response(data, metrics_hint=args.metrics)
            for rec in recs:
                rec["day"] = params['start'][:10]
            _ensure_hostname_column(recs)

            if not recs:
                days_no_data.append(params['start'][:10])
                cur = nxt
                continue

            # Scrittura in base al formato
            if args.format == "csv":
                if header_keys is None:
                    # deduci chiavi dalla prima batch
                    keys: List[str] = []
                    seen = set()
                    preferred = ["day", "hostname", "url", "timestamp", "cpcode", "interval"]
                    for rec in recs:
                        for k in rec.keys():
                            if k not in seen:
                                seen.add(k)
                                keys.append(k)
                    def srt(k: str):
                        return (0, preferred.index(k)) if k in preferred else (1, k)
                    keys.sort(key=srt)
                    _write_csv_header(args.out, keys, args.csv_delimiter)
                    header_keys = keys
                _append_csv_rows(args.out, header_keys, recs, args.csv_delimiter)

            elif args.format == "xlsx":
                if header_keys is None:
                    keys = []
                    seen = set()
                    preferred = ["day", "hostname", "url", "timestamp", "cpcode", "interval"]
                    for rec in recs:
                        for k in rec.keys():
                            if k not in seen:
                                seen.add(k)
                                keys.append(k)
                    def srt(k: str):
                        return (0, preferred.index(k)) if k in preferred else (1, k)
                    keys.sort(key=srt)
                    _write_xlsx_header(args.out, keys, args.sheet_name)
                    header_keys = keys
                _append_xlsx_rows(args.out, header_keys, recs, args.sheet_name)

            else:  # json
                json_aggregate.append(recs)

            total_rows += len(recs)
            if args.verbose:
                eprint(f"  {params['start'][:10]}: +{len(recs)} righe (totale {total_rows})")
                sys.stderr.flush()

            cur = nxt

        if days_no_data:
            sys.stderr.write("Giorni senza dati: " + ", ".join(days_no_data) + "")
            sys.stderr.flush()

        if args.format == "json":
            if args.out:
                with open(args.out, "w", encoding="utf-8") as f:
                    json.dump(json_aggregate, f, ensure_ascii=False, indent=2 if args.pretty else None)
            print(json.dumps(json_aggregate, ensure_ascii=False, indent=2 if args.pretty else None))
        else:
            # Retry su giorni senza dati (solo per CSV/XLSX)
            if days_no_data:
                sys.stderr.write("Giorni senza dati: " + ", ".join(days_no_data) + ""); sys.stderr.flush()
                still_missing: List[str] = []
                for d in days_no_data:
                    for attempt in range(1, args.retry_no_data + 1):
                        print(f"Retry {d} ({attempt}/{args.retry_no_data})...", file=sys.stderr, flush=True)
                        dstart, dend = _day_to_range(d)
                        params = {"start": dstart, "end": dend, "interval": "DAY"}
                        body = {"objectType": args.object_type, "objectIds": args.object_ids, "metrics": args.metrics, "limit": 25000}
                        try:
                            r = session.post(base_url, params=params, json=body, headers={"Content-Type": "application/json"}, timeout=args.timeout)
                        except requests.exceptions.RequestException as e:
                            eprint(f"  [WARN] Retry fallito per {d}: {e}"); sys.stderr.flush(); time.sleep(args.retry_wait); continue
                        if r.status_code >= 400:
                            eprint(f"  [WARN] HTTP {r.status_code} su {d}"); sys.stderr.flush(); time.sleep(args.retry_wait); continue
                        if not r.content or r.text.strip() == "":
                            time.sleep(args.retry_wait); continue
                        try:
                            data = r.json()
                        except Exception:
                            time.sleep(args.retry_wait); continue
                        recs = _records_from_response(data, metrics_hint=args.metrics)
                        for rec in recs:
                            rec["day"] = d
                        _ensure_hostname_column(recs)
                        if not recs:
                            time.sleep(args.retry_wait); continue
                        if args.format == "csv":
                            if header_keys is None:
                                keys: List[str] = []
                                seen = set()
                                preferred = ["day", "hostname", "url", "timestamp", "cpcode", "interval"]
                                for rec in recs:
                                    for k in rec.keys():
                                        if k not in seen:
                                            seen.add(k); keys.append(k)
                                def srt(k: str):
                                    return (0, preferred.index(k)) if k in preferred else (1, k)
                                keys.sort(key=srt)
                                _write_csv_header(args.out, keys, args.csv_delimiter)
                                header_keys = keys
                            _append_csv_rows(args.out, header_keys, recs, args.csv_delimiter)
                        elif args.format == "xlsx":
                            if header_keys is None:
                                keys = []
                                seen = set()
                                preferred = ["day", "hostname", "url", "timestamp", "cpcode", "interval"]
                                for rec in recs:
                                    for k in rec.keys():
                                        if k not in seen:
                                            seen.add(k); keys.append(k)
                                def srt(k: str):
                                    return (0, preferred.index(k)) if k in preferred else (1, k)
                                keys.sort(key=srt)
                                _write_xlsx_header(args.out, keys, args.sheet_name)
                                header_keys = keys
                            _append_xlsx_rows(args.out, header_keys, recs, args.sheet_name)
                        total_rows += len(recs)
                        eprint(f"  [OK] Recuperato {len(recs)} righe per {d}"); sys.stderr.flush()
                        break
                    else:
                        still_missing.append(d)
                days_no_data = still_missing
                if days_no_data:
                    sys.stderr.write("Ancora senza dati: " + ", ".join(days_no_data) + ""); sys.stderr.flush()
                else:
                    eprint("Tutti i giorni senza dati sono stati recuperati."); sys.stderr.flush()

            eprint(f"[OK] MAX completata. Righe totali: {total_rows}"); sys.stderr.flush()
        sys.exit(0)

    # --- Percorso standard (non MAX) ---
    params = {"start": args.start, "end": args.end, "interval": args.interval}
    body = {
        "objectType": args.object_type,
        "objectIds": args.object_ids,
        "metrics": args.metrics,
        "limit": int(args.limit) if str(args.limit).isdigit() else 5000,
    }

    if args.verbose or args.dry_run:
        eprint("Invio POST a:")
        eprint(f"  URL: {base_url}")
        eprint(f"  Query: {json.dumps(params, ensure_ascii=False)}")
        eprint(f"  Body: {json.dumps(body, ensure_ascii=False)}")
        sys.stderr.flush()

    if args.dry_run:
        eprint("--dry-run attivo: nessuna richiesta inviata.")
        sys.stderr.flush()
        sys.exit(0)

    try:
        r = session.post(base_url, params=params, json=body, headers={"Content-Type": "application/json"}, timeout=args.timeout)
    except requests.exceptions.Timeout:
        eprint(f"[ERRORE] Timeout dopo {args.timeout}s. Verifica connettività verso {creds['host']} o riduci il range.")
        sys.stderr.flush()
        sys.exit(2)
    except requests.exceptions.RequestException as e:
        eprint(f"[ERRORE] Richiesta fallita: {e}")
        sys.stderr.flush()
        sys.exit(2)

    if args.log_headers:
        for k, v in r.headers.items():
            eprint(f"{k}: {v}")
        sys.stderr.flush()

    if r.status_code >= 400:
        try:
            payload = r.json()
        except Exception:
            payload = {"error": r.text}
        eprint(json.dumps(payload, ensure_ascii=False, indent=2))
        sys.stderr.flush()
        sys.exit(1)

    if not r.content or r.text.strip() == "":
        eprint("(Nessun contenuto)")
        sys.stderr.flush()
        if args.format == "csv":
            if args.out:
                open(args.out, "w", encoding="utf-8-sig").close()
        elif args.format == "xlsx":
            # crea file vuoto con solo header se possibile (senza dati non conosciamo le chiavi)
            if args.out:
                _write_xlsx_header(args.out, ["day"], args.sheet_name)
        else:
            print("{}")
        sys.exit(0)

    try:
        data = r.json()
    except Exception:
        if args.format == "csv":
            eprint("[ERRORE] Payload non-JSON: impossibile CSV.")
            sys.stderr.flush()
            sys.exit(1)
        elif args.format == "xlsx":
            eprint("[ERRORE] Payload non-JSON: impossibile XLSX.")
            sys.stderr.flush()
            sys.exit(1)
        else:
            if args.out:
                with open(args.out, "w", encoding="utf-8") as f:
                    f.write(r.text)
            else:
                sys.stdout.write(r.text)
            sys.exit(0)

    recs = _records_from_response(data, metrics_hint=args.metrics)

    if args.format == "csv":
        if recs:
            keys: List[str] = []
            seen = set()
            preferred = ["hostname", "url", "timestamp", "cpcode", "interval"]
            for rec in recs:
                for k in rec.keys():
                    if k not in seen:
                        seen.add(k)
                        keys.append(k)
            def srt(k: str):
                return (0, preferred.index(k)) if k in preferred else (1, k)
            keys.sort(key=srt)
            _write_csv_header(args.out, keys, args.csv_delimiter)
            _append_csv_rows(args.out, keys, recs, args.csv_delimiter)
        else:
            if args.out:
                open(args.out, "w", encoding="utf-8-sig").close()
        sys.exit(0)

    if args.format == "xlsx":
        if not args.out:
            args.out = "report.xlsx"
        if recs:
            keys = []
            seen = set()
            preferred = ["hostname", "url", "timestamp", "cpcode", "interval"]
            for rec in recs:
                for k in rec.keys():
                    if k not in seen:
                        seen.add(k)
                        keys.append(k)
            def srt(k: str):
                return (0, preferred.index(k)) if k in preferred else (1, k)
            keys.sort(key=srt)
            _write_xlsx_header(args.out, keys, args.sheet_name)
            _append_xlsx_rows(args.out, keys, recs, args.sheet_name)
        else:
            _write_xlsx_header(args.out, ["day"], args.sheet_name)
        sys.exit(0)

    # JSON
    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2 if args.pretty else None)
    print(json.dumps(data, ensure_ascii=False, indent=2 if args.pretty else None))


if __name__ == "__main__":
    main()
